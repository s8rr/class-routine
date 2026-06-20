"""
Extract the DCSE class routine out of the official term spreadsheet
(routine.xlsx) into data/routine_data.json and data/img_data.json.

Sheet layout (same on every weekday tab):
  Row 5    : time-slot headers, e.g. "8.30 AM- 8.55 AM"
  Col B    : "Sem" - the semester+section label, e.g. "1A", "5C", "8E"
  Col C..  : one cell per scheduled class, formatted "CODE/TEACHER/ROOM".
             A class that runs across several time-slots is stored as a
             merged cell spanning those columns.

A semester+section can spill onto a second spreadsheet row when it has more
classes than fit visually in one row (e.g. two lab batches running at the
same time) - in that case the "Sem" cell itself is merged down across both
rows, so its value only shows up on the first of the two. We forward-fill
the current section as we walk down the sheet to handle that.
"""

import os
import re
import json
import openpyxl

XLSX_PATH = "routine.xlsx"
JSON_PATH = "data/routine_data.json"
IMG_JSON_PATH = "data/img_data.json"

# Worksheet tab name (lowercased) -> canonical day name used in the JSON output.
DAY_SHEET_MAP = {
    "saturday": "Saturday",
    "sunday": "Sunday",
    "monday": "Monday",
    "tuesday": "Tuesday",
    "wednesday": "Wednesday",
}
ALL_DAYS = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

SEM_COL = 2          # column B holds the "Sem" (semester+section) label
FIRST_DATA_COL = 3   # column C is where the time-slot grid starts
HEADER_ROW = 5
FIRST_DATA_ROW = 6
MAX_SCAN_ROWS = 400   # generous ceiling; we stop early once the table clearly ends

CLASS_RE = re.compile(r'^\s*([^/]+?)\s*/\s*(.+?)\s*/\s*([^/]+?)\s*$')
TIME_RE = re.compile(
    r'(\d{1,2})[.:](\d{2})\s*(?:[ap]\.?\s*m\.?)?\s*-\s*(\d{1,2})[.:](\d{2})\s*(?:[ap]\.?\s*m\.?)?',
    re.IGNORECASE,
)


def resolve_12h(hour12, minute, floor_minutes):
    """Pick whichever am/pm reading keeps the schedule moving forward.

    The source header row has at least one mislabeled am/pm (e.g. an
    11:25 AM slot tagged "pm"), so instead of trusting the text we infer
    the correct half of the day from what time slot came right before it.
    """
    candidates = []
    for am_pm in ("AM", "PM"):
        h24 = hour12 % 12
        if am_pm == "PM":
            h24 += 12
        candidates.append((h24 * 60 + minute, am_pm))
    candidates.sort()

    if floor_minutes <= 0:
        chosen = next((c for c in candidates if c[1] == "AM"), candidates[0])
    else:
        forward = [c for c in candidates if c[0] >= floor_minutes]
        chosen = min(forward) if forward else min(candidates)

    total, am_pm = chosen
    h12 = (total // 60) % 12 or 12
    label = f"{h12:02d}:{total % 60:02d} {am_pm}"
    return total, label


def build_time_slots(ws):
    """Read the header row once and return a list of
    {col, start, end, end_minutes} for every time-slot column found."""
    slots = []
    prev_end_minutes = -1
    col = FIRST_DATA_COL
    empty_streak = 0
    while empty_streak < 5:
        header = ws.cell(row=HEADER_ROW, column=col).value
        if header is None or not str(header).strip():
            empty_streak += 1
            col += 1
            continue
        empty_streak = 0
        match = TIME_RE.search(str(header))
        if not match:
            col += 1
            continue
        sh, sm, eh, em = (int(x) for x in match.groups())
        start_minutes, start_label = resolve_12h(sh, sm, prev_end_minutes)
        end_minutes, end_label = resolve_12h(eh, em, start_minutes)
        slots.append({"col": col, "start": start_label, "end": end_label, "end_minutes": end_minutes})
        prev_end_minutes = end_minutes
        col += 1
    return slots


def merge_lookup(ws):
    """Map every (row, col) that is the *top-left* of a merged range to that
    range's last column, so we know how many time-slots a class spans."""
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        lookup[(merged_range.min_row, merged_range.min_col)] = merged_range.max_col
    return lookup


def parse_day_sheet(ws, time_slots):
    """Walk one weekday worksheet and return {section: [class, ...]}."""
    col_to_slot = {slot["col"]: slot for slot in time_slots}
    last_col = time_slots[-1]["col"] if time_slots else FIRST_DATA_COL
    merges = merge_lookup(ws)

    routine_by_section = {}
    current_section = None

    for row in range(FIRST_DATA_ROW, MAX_SCAN_ROWS):
        sem_value = ws.cell(row=row, column=SEM_COL).value
        row_has_any_value = any(
            ws.cell(row=row, column=c).value is not None
            for c in range(SEM_COL, last_col + 1)
        )

        if sem_value is None and not row_has_any_value:
            if current_section is None:
                continue
            # A run of fully blank rows means the table is over.
            blank_ahead = all(
                ws.cell(row=row + i, column=c).value is None
                for i in range(3)
                for c in range(SEM_COL, last_col + 1)
            )
            if blank_ahead:
                break
            continue

        if sem_value is not None:
            current_section = str(sem_value).strip().upper()
        if not current_section:
            continue

        routine_by_section.setdefault(current_section, [])

        for col in range(FIRST_DATA_COL, last_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                continue

            start_slot = col_to_slot.get(col)
            if start_slot is None:
                continue
            end_col = merges.get((row, col), col)
            end_slot = col_to_slot.get(end_col, start_slot)

            match = CLASS_RE.match(str(cell.value))
            if not match:
                continue
            subject, teacher, room = (g.strip() for g in match.groups())

            routine_by_section[current_section].append({
                "time": f"{start_slot['start']} - {end_slot['end']}",
                "subject": subject,
                "teacher": teacher,
                "room": room,
                "_sort": start_slot["end_minutes"],
            })

    for classes in routine_by_section.values():
        classes.sort(key=lambda c: c["_sort"])
        for c in classes:
            del c["_sort"]

    return routine_by_section


def semester_of(section):
    match = re.match(r'(\d+)', section)
    return match.group(1) if match else section


def parse_routine_xlsx():
    print("Starting XLSX data extraction pipeline...")

    if not os.path.exists(XLSX_PATH):
        print(f"Error: {XLSX_PATH} not found. Nothing to extract.")
        return

    # Preserve hand-managed fields (the 'overright' approval flag and any
    # custom holiday/exam exceptions) across re-extractions.
    existing = {}
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, "r") as f:
            try:
                existing = json.load(f)
            except json.JSONDecodeError:
                existing = {}

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    weekly_routine = {}
    all_sections = set()

    for sheet_name in wb.sheetnames:
        day = DAY_SHEET_MAP.get(sheet_name.strip().lower())
        if day is None:
            print(f"Skipping unrecognized sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        time_slots = build_time_slots(ws)
        day_routine = parse_day_sheet(ws, time_slots)

        for section, classes in day_routine.items():
            all_sections.add(section)
            weekly_routine.setdefault(section, {d: [] for d in ALL_DAYS})
            weekly_routine[section][day] = classes

    for section in all_sections:
        for d in ALL_DAYS:
            weekly_routine[section].setdefault(d, [])

    sections_by_semester = {}
    for section in sorted(all_sections, key=lambda s: (semester_of(s), s)):
        sections_by_semester.setdefault(semester_of(section), []).append(section)

    semesters = sorted(sections_by_semester.keys(), key=lambda s: int(s))

    data_store = {
        "overright": existing.get("overright", False),
        "semesters": semesters,
        "sections_by_semester": sections_by_semester,
        "sections": sorted(all_sections, key=lambda s: (semester_of(s), s)),
        "weekly_routine": weekly_routine,
        "exceptions": existing.get("exceptions", {}),
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w") as f:
        json.dump(data_store, f, indent=2)

    # img_data.json feeds generate_images.py; same shape works fine there.
    with open(IMG_JSON_PATH, "w") as f:
        json.dump(data_store, f, indent=2)

    total_classes = sum(len(c) for sec in weekly_routine.values() for c in sec.values())
    print(
        f"Extracted {len(all_sections)} sections across {len(semesters)} semesters "
        f"({total_classes} scheduled classes) -> {JSON_PATH} & {IMG_JSON_PATH}"
    )
    print("Routine extraction pipeline completed successfully. Data synchronized.")


if __name__ == "__main__":
    parse_routine_xlsx()
